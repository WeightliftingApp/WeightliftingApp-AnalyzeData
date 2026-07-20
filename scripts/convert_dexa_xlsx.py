#!/usr/bin/env python3
"""Extract DEXA measurements from Weight Log.xlsx into analysis-ready CSVs."""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_INPUT = Path(
    "/Users/chappyasel/Library/Mobile Documents/com~apple~CloudDocs/Spreadsheets/Weight Log.xlsx"
)
DEFAULT_TOTAL_OUTPUT = Path(__file__).resolve().parents[1] / "data" / "dexa.csv"
DEFAULT_REGION_OUTPUT = (
    Path(__file__).resolve().parents[1] / "data" / "dexa_regions.csv"
)


def _number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _rounded(value: float | None, digits: int = 4) -> float | None:
    return None if value is None else round(value, digits)


def _find_total_header_row(sheet) -> int:
    for row in range(1, sheet.max_row):
        if sheet.cell(row, 1).value == "DEXA" and sheet.cell(row + 1, 2).value == "Weight":
            return row + 1
    raise ValueError("Unable to find the DEXA totals table in the BF% sheet")


def _find_height_inches(sheet) -> float:
    """Find the numeric height stored below the workbook's Height label."""
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(row, column).value
            if isinstance(value, str) and value.strip().lower() == "height":
                for value_row in range(row + 1, min(row + 4, sheet.max_row + 1)):
                    height_in = _number(sheet.cell(value_row, column).value)
                    if height_in is not None:
                        return height_in
    raise ValueError("Unable to find a numeric height below the Height label")


def extract_totals(sheet) -> list[dict[str, Any]]:
    header_row = _find_total_header_row(sheet)
    height_in = _find_height_inches(sheet)

    height_m = height_in * 0.0254
    records: list[dict[str, Any]] = []

    for row in range(header_row + 1, sheet.max_row + 1):
        scan_date = sheet.cell(row, 1).value
        if not isinstance(scan_date, (date, datetime)):
            break

        weight = _number(sheet.cell(row, 2).value)
        lean_soft_tissue = _number(sheet.cell(row, 3).value)
        fat_mass = _number(sheet.cell(row, 4).value)
        bone_mineral_content = _number(sheet.cell(row, 5).value)
        if weight is None or fat_mass is None:
            continue

        fat_free_mass = weight - fat_mass
        ffmi = (fat_free_mass * 0.45359237) / height_m**2
        normalized_ffmi = ffmi + 6.1 * (1.8 - height_m)
        bmi = (weight * 0.45359237) / height_m**2

        records.append(
            {
                "date": scan_date.strftime("%Y-%m-%d"),
                "weight_lb": _rounded(weight, 1),
                "lean_soft_tissue_lb": _rounded(lean_soft_tissue, 1),
                "fat_mass_lb": _rounded(fat_mass, 1),
                "bone_mineral_content_lb": _rounded(bone_mineral_content, 1),
                "body_fat_pct": _rounded(100 * fat_mass / weight, 2),
                "fat_free_mass_lb": _rounded(fat_free_mass, 1),
                "ffmi": _rounded(ffmi, 2),
                "normalized_ffmi": _rounded(normalized_ffmi, 2),
                "bmi": _rounded(bmi, 2),
                "height_in": _rounded(height_in, 1),
                "notes": sheet.cell(row, 14).value or "",
            }
        )

    if not records:
        raise ValueError("The DEXA totals table did not contain any scans")
    return records


def _find_regional_date_row(sheet) -> int:
    total_header_row = _find_total_header_row(sheet)
    for row in range(total_header_row + 1, sheet.max_row):
        if sheet.cell(row, 1).value == "DEXA" and isinstance(
            sheet.cell(row + 1, 3).value, (date, datetime)
        ):
            return row + 1
    raise ValueError("Unable to find the DEXA regional table in the BF% sheet")


def extract_regions(sheet) -> list[dict[str, Any]]:
    date_row = _find_regional_date_row(sheet)
    scan_columns = {
        column: sheet.cell(date_row, column).value
        for column in range(3, sheet.max_column + 1)
        if isinstance(sheet.cell(date_row, column).value, (date, datetime))
    }
    records: list[dict[str, Any]] = []

    for row in range(date_row + 1, sheet.max_row - 2):
        region = sheet.cell(row, 1).value
        if not isinstance(region, str) or sheet.cell(row, 2).value != "Fat":
            continue
        if sheet.cell(row + 1, 2).value != "Lean":
            continue

        for column, scan_date in scan_columns.items():
            fat_mass = _number(sheet.cell(row, column).value)
            lean_soft_tissue = _number(sheet.cell(row + 1, column).value)
            bone_mineral_content = (
                _number(sheet.cell(row + 2, column).value)
                if sheet.cell(row + 2, 2).value == "BMC"
                else None
            )
            if fat_mass is None and lean_soft_tissue is None:
                continue

            regional_total = sum(
                value
                for value in (fat_mass, lean_soft_tissue, bone_mineral_content)
                if value is not None
            )
            body_fat_pct = (
                100 * fat_mass / regional_total
                if fat_mass is not None
                and bone_mineral_content is not None
                and regional_total > 0
                else None
            )
            records.append(
                {
                    "date": scan_date.strftime("%Y-%m-%d"),
                    "region": " ".join(region.split()),
                    "fat_mass_lb": _rounded(fat_mass, 1),
                    "lean_soft_tissue_lb": _rounded(lean_soft_tissue, 1),
                    "bone_mineral_content_lb": _rounded(
                        bone_mineral_content, 1
                    ),
                    "body_fat_pct": _rounded(body_fat_pct, 2),
                }
            )

    if not records:
        raise ValueError("The DEXA regional table did not contain any measurements")
    return records


def write_csv(path: Path, records: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=records[0].keys())
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--totals-output", type=Path, default=DEFAULT_TOTAL_OUTPUT)
    parser.add_argument("--regions-output", type=Path, default=DEFAULT_REGION_OUTPUT)
    args = parser.parse_args()

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    sheet = workbook["BF%"]
    totals = extract_totals(sheet)
    regions = extract_regions(sheet)
    write_csv(args.totals_output, totals)
    write_csv(args.regions_output, regions)

    latest = totals[-1]
    print(f"Extracted {len(totals)} DEXA scans -> {args.totals_output}")
    print(f"Extracted {len(regions)} regional measurements -> {args.regions_output}")
    print(
        f"Latest: {latest['date']} — {latest['weight_lb']:.1f} lb, "
        f"{latest['body_fat_pct']:.2f}% body fat, FFMI {latest['ffmi']:.2f}"
    )


if __name__ == "__main__":
    main()
