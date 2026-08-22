#!/usr/bin/env python3
"""Generate the post-ready flat-bench Pareto frontier update card."""

from datetime import datetime, timedelta
from pathlib import Path

from matplotlib.lines import Line2D
import openpyxl
import pandas as pd

from analysis_utils import build_trimmed_trailing_bodyweight, mark_pareto_frontier
from chart_style import (
    MONO_FONT,
    PALETTE,
    PARETO_AXES,
    PARETO_FRAME,
    add_footer,
    add_header,
    chart_canvas,
    save_chart,
    style_axes,
)
from schema import WLD

REPO_ROOT = Path(__file__).resolve().parents[1]
WLD_PATH = Path.home() / "Downloads/WeightliftingAppData.wld"
WEIGHT_LOG_PATH = (
    Path.home()
    / "Library/Mobile Documents/com~apple~CloudDocs/Spreadsheets/Weight Log.xlsx"
)
OUTPUT_PATH = REPO_ROOT / "outputs/bench-strength-eval-update-2026-08-21.png"
UPDATE_DATE = pd.Timestamp("2026-08-21")

SKIP_SHEETS = {"Agg", "BF%", "FFMI", "Sheet1"}
COL_WEEK_OF = 3
DAY_COLUMNS = range(4, 11)


def load_daily_weights(path: Path) -> pd.DataFrame:
    """Read daily weigh-ins from every period sheet in the Weight Log."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, object]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        for row in rows[2:]:
            if len(row) <= 10 or not isinstance(row[COL_WEEK_OF], datetime):
                continue
            week_of = row[COL_WEEK_OF]
            for day_offset, column in enumerate(DAY_COLUMNS):
                weight = row[column]
                if isinstance(weight, (int, float)) and weight > 0:
                    records.append(
                        {
                            "date": week_of + timedelta(days=day_offset),
                            "weight": float(weight),
                            "sheet": sheet_name,
                        }
                    )
    workbook.close()
    return (
        pd.DataFrame(records)
        .sort_values(["date", "sheet"])
        .drop_duplicates("date", keep="last")
    )


def load_flat_bench_attempts(path: Path) -> pd.DataFrame:
    """Read every valid flat barbell bench estimated-1RM attempt."""
    wld = WLD(file_path=str(path))
    records: list[dict[str, object]] = []
    for workout in wld.workouts:
        for exercise in workout.exercises:
            if exercise.displayName() != "Flat Barbell Bench Press":
                continue
            for set_order, set_data in enumerate(exercise.sets):
                if set_data.oneRM is None or set_data.oneRM <= 0:
                    continue
                records.append(
                    {
                        "date": pd.Timestamp(workout.date).normalize(),
                        "workout_id": workout.uuid,
                        "set_order": set_order,
                        "weight": set_data.weight,
                        "reps": set_data.reps,
                        "one_rm": set_data.oneRM,
                    }
                )
    return pd.DataFrame(records)


def unique_frontier(attempts: pd.DataFrame) -> pd.DataFrame:
    return (
        attempts[attempts["is_pareto"]]
        .sort_values(["bodyweight", "one_rm"])
        .drop_duplicates(["bodyweight", "one_rm"])
        .reset_index(drop=True)
    )


def render_frontier_update(
    attempts: pd.DataFrame, output_path: Path, update_date: pd.Timestamp = UPDATE_DATE,
) -> dict[str, object]:
    """Render a bench Pareto update from attempts with matched bodyweights."""
    previous_attempts = mark_pareto_frontier(attempts[attempts["date"] < update_date])
    current_attempts = mark_pareto_frontier(attempts[attempts["date"] <= update_date])
    previous_frontier = unique_frontier(previous_attempts)
    current_frontier = unique_frontier(current_attempts)

    today = current_attempts[current_attempts["date"] == update_date]
    if today.empty:
        raise RuntimeError(
            f"No matched flat-bench attempts found on {update_date:%Y-%m-%d}"
        )
    new_point = today.loc[today["one_rm"].idxmax()]
    if not bool(new_point["is_pareto"]):
        raise RuntimeError(
            "Today's best flat-bench attempt is not on the corrected frontier"
        )

    old_coordinates = set(
        previous_frontier[["bodyweight", "one_rm"]].itertuples(index=False, name=None)
    )
    added = current_frontier[
        ~current_frontier.apply(
            lambda row: (row["bodyweight"], row["one_rm"]) in old_coordinates, axis=1,
        )
    ]
    if len(added) != 1:
        raise RuntimeError(f"Expected one added frontier point, found {len(added)}")

    new_index = current_frontier.index[
        (current_frontier["bodyweight"] == new_point["bodyweight"])
        & (current_frontier["one_rm"] == new_point["one_rm"])
    ][0]
    segment_start = max(0, new_index - 1)
    segment_end = min(len(current_frontier) - 1, new_index + 1)
    red_segment = current_frontier.iloc[segment_start : segment_end + 1]

    with chart_canvas(PARETO_FRAME) as (fig, ax):
        add_header(
            fig,
            PARETO_FRAME,
            "BENCH PRESS",
            "Pareto frontier update · today established a new checkpoint",
            (
                f"EVAL WINDOW  {attempts['date'].min():%Y.%m}–{attempts['date'].max():%Y.%m}",
                f"{attempts['workout_id'].nunique():,} WORKOUTS  /  {len(attempts):,} ATTEMPTS",
            ),
        )

        ax.scatter(
            current_attempts["bodyweight"],
            current_attempts["one_rm"],
            s=27,
            color=PALETTE.checkpoint,
            alpha=0.42,
            linewidth=0,
            rasterized=True,
            zorder=1,
        )
        ax.plot(
            previous_frontier["bodyweight"],
            previous_frontier["one_rm"],
            color=PALETTE.frontier,
            linewidth=3.0,
            linestyle=(0, (5, 4)),
            alpha=0.78,
            zorder=2,
        )
        ax.scatter(
            previous_frontier["bodyweight"],
            previous_frontier["one_rm"],
            s=70,
            color=PALETTE.frontier,
            edgecolor=PALETTE.panel,
            linewidth=1.3,
            alpha=0.8,
            zorder=3,
        )
        ax.plot(
            red_segment["bodyweight"],
            red_segment["one_rm"],
            color=PALETTE.advance,
            linewidth=4.6,
            solid_capstyle="round",
            zorder=5,
        )
        ax.scatter(
            [new_point["bodyweight"]],
            [new_point["one_rm"]],
            s=190,
            color=PALETTE.advance,
            edgecolor=PALETTE.panel,
            linewidth=2.2,
            zorder=6,
        )
        ax.annotate(
            f'NEW FRONTIER  ·  {int(new_point["weight"])}×{int(new_point["reps"])}\n'
            f'{new_point["one_rm"]:.0f} 1RMe @ {new_point["bodyweight"]:.1f} BW',
            (new_point["bodyweight"], new_point["one_rm"]),
            xytext=(0, 28),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=10,
            family=MONO_FONT,
            color=PALETTE.advance_ink,
            fontweight="bold",
            zorder=7,
            arrowprops={"arrowstyle": "-", "color": PALETTE.advance, "linewidth": 1.8,},
        )

        ax.text(
            0.015,
            0.975,
            "BETTER  ↖",
            transform=ax.transAxes,
            fontsize=8.5,
            family=MONO_FONT,
            color=PALETTE.muted,
            ha="left",
            va="top",
        )
        ax.text(
            0.985,
            0.975,
            f"MAX SCORE  {current_frontier['one_rm'].max():.0f}",
            transform=ax.transAxes,
            fontsize=8.5,
            family=MONO_FONT,
            color=PALETTE.muted,
            ha="right",
            va="top",
        )

        ax.set_xlabel(
            "COST PROXY · BODYWEIGHT (LB)",
            fontsize=10.5,
            family=MONO_FONT,
            labelpad=12,
        )
        ax.set_ylabel(
            "CAPABILITY SCORE · 1RME (LB)",
            fontsize=10.5,
            family=MONO_FONT,
            labelpad=12,
        )
        ax.set_xlim(attempts["bodyweight"].min() - 2, attempts["bodyweight"].max() + 2)
        ax.set_ylim(300, 525)
        style_axes(ax, PARETO_AXES)

        handles = [
            Line2D(
                [0],
                [0],
                color=PALETTE.frontier,
                linewidth=3,
                linestyle=(0, (5, 4)),
                marker="o",
                markersize=7,
                label=f"PREVIOUS FRONTIER  n={len(previous_frontier)}",
            ),
            Line2D(
                [0],
                [0],
                color=PALETTE.advance,
                linewidth=4,
                marker="o",
                markersize=9,
                label="TODAY'S ADVANCE",
            ),
        ]
        ax.legend(
            handles=handles,
            loc="lower right",
            frameon=False,
            prop={"family": MONO_FONT, "size": 8.5},
            labelcolor=PALETTE.muted,
            handletextpad=0.7,
        )

        add_footer(
            fig,
            PARETO_FRAME,
            "BODYWEIGHT: TRAILING 7D TRIMMED AVG · DROP HIGH + LOW, AVG MIDDLE 5",
            "FRONTIER: NON-DOMINATED  ·  MAXIMIZE PERFORMANCE / COST  ↖",
            right_color=PALETTE.advance,
        )
        save_chart(fig, output_path, dpi=PARETO_FRAME.dpi)

    return {
        "output": str(output_path),
        "attempts": len(attempts),
        "workouts": int(attempts["workout_id"].nunique()),
        "previous_frontier_points": len(previous_frontier),
        "current_frontier_points": len(current_frontier),
        "new_bodyweight": round(float(new_point["bodyweight"]), 2),
        "new_one_rm": int(new_point["one_rm"]),
        "set": f'{int(new_point["weight"])}x{int(new_point["reps"])}',
    }


def build_chart(
    wld_path: Path = WLD_PATH,
    weight_log_path: Path = WEIGHT_LOG_PATH,
    output_path: Path = OUTPUT_PATH,
    update_date: pd.Timestamp = UPDATE_DATE,
) -> dict[str, object]:
    """Load source files and write the post-ready flat-bench update card."""
    daily_weights = load_daily_weights(weight_log_path)
    trailing_weights = build_trimmed_trailing_bodyweight(daily_weights)
    attempts = load_flat_bench_attempts(wld_path).merge(
        trailing_weights, on="date", how="left"
    )
    attempts = attempts.dropna(subset=["bodyweight"]).copy()
    return render_frontier_update(attempts, output_path, update_date)


if __name__ == "__main__":
    print(build_chart())
